import re
import typing
from abc import ABC, abstractmethod
import dataclasses
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook


@dataclasses.dataclass
class ReceipLine:
    number: int
    name: str
    price: float
    quantity: float
    cost: float


@dataclasses.dataclass
class Receip:
    date: datetime
    link: str
    currency: str
    vendor: str
    items: typing.List[ReceipLine] = dataclasses.field(default_factory=list)
    extraData: typing.Dict[str, str] = dataclasses.field(default_factory=dict)

    @property
    def amount(self) -> float:
        return sum(i.cost for i in self.items)

    @property
    def volume(self) -> int:
        return len(self.items)

    def save_virtual_xlsx(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['vendor', self.vendor])
        ws.append(['currency', self.currency])
        ws.append(['date', self.date])
        ws.append(['link', self.link])
        for kv in self.extraData.items():
            ws.append(kv)
        ws.append([])
        fields = [field.name for field in dataclasses.fields(ReceipLine)]
        ws.append(fields + ['date'])
        date = self.date.strftime('%d.%m.%Y')
        for item in self.items:
            ws.append([getattr(item, i) for i in fields] + [date])

        cell_widths = {}
        for column in ws:
            for cell in column:
                if cell.value:
                    if str(cell.value)[0] == "=":
                        continue
                    cell_widths[cell.column] = max(
                        (cell_widths.get(cell.column, 0), len(str(cell.value)))
                    )
        for col, column_width in cell_widths.items():
            column_width = str(column_width + 2)
            ws.column_dimensions[get_column_letter(col)].width = column_width
        return save_virtual_workbook(wb)


class ReceipParser(ABC):
    templateRegStr: str = ""
    name: str = 'abstract'

    def __init__(self):
        self.templateRegEx = re.compile(self.templateRegStr)

    @staticmethod
    def get_all_vendors():
        return {i.name for i in ReceipParser.parsers()}

    @staticmethod
    def parsers():
        classes = set()
        work = [ReceipParser]
        while work:
            parent = work.pop()
            for child in parent.__subclasses__():
                if child not in classes:
                    classes.add(child)
                    work.append(child)
        return classes

    @abstractmethod
    async def parser(self, link: str) -> Receip:
        pass

    @abstractmethod
    async def extract_link(self, message: str) -> str:
        pass

    async def parse(self, message: str) -> Receip:
        return await self.parser(await self.extract_link(message))
